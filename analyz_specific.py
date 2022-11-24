import openpyxl
from openpyxl.utils import get_column_letter
import re

path_to_file = 'data.xlsx'

search_text = 'искусственный интеллект'
search_text = search_text
print('Ищем:', search_text)

wb = openpyxl.load_workbook(path_to_file)
sheets_list = wb.sheetnames
sheet_active = wb[sheets_list[0]]
row_max = sheet_active.max_row

column_max = sheet_active.max_column

print('В файле:', path_to_file, '\n Cтолбцов:', row_max, '\n Колонок:', column_max)

row_min = 1
column_min = 1

while column_min <= column_max:
    row_min_min = row_min
    row_max_max = row_max
    while row_min_min <= row_max_max:
        row_min_min = str(row_min_min)

        word_column = get_column_letter(column_min)
        word_column = str(word_column)
        word_cell = word_column + row_min_min

        data_from_cell = sheet_active[word_cell].value
        data_from_cell = str(data_from_cell)
        #print(data_from_cell)
        regular = search_text
        result = re.findall(regular, data_from_cell)
        if len(result) > 0:
            print('Нашли в ячейке:', word_cell)
        row_min_min = int(row_min_min)
        row_min_min = row_min_min + 1
    column_min = column_min + 1
